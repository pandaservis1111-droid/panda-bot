from aiogram import Router, F
from aiogram.types import Message, KeyboardButton, ReplyKeyboardMarkup, ReplyKeyboardRemove, FSInputFile, InlineKeyboardMarkup, InlineKeyboardButton, CallbackQuery
from aiogram.filters import CommandStart
from aiogram.fsm.state import State, StatesGroup
from aiogram.fsm.context import FSMContext
from openpyxl import Workbook, load_workbook
from pathlib import Path
from datetime import datetime
import os

router = Router()
def parse_ids(s: str):
    ids = []
    for x in s.split(","):
        x = x.strip()
        if x:
            try:
                ids.append(int(x))
            except ValueError:
                print(f"⚠️ Невалидный ID: {x}")
    return ids

ADMIN_IDS = parse_ids(os.getenv("ADMIN_IDS", ""))
GROUP_IDS = parse_ids(os.getenv("GROUP_IDS", ""))
EXCEL_FILE = "users.xlsx"

LANG = {
    "🇺🇿 O‘zbek": {
        "welcome": "🐼 <b>Panda Servisga xush kelibsiz!</b>\n\nRo‘yxatdan o‘tish uchun quyidagi ma’lumotlarni kiriting 👇",
        "choose_language": "Tilni tanlang 👇",
        "enter_name": "🧾 Ismingizni kiriting:",
        "enter_phone": "📱 Telefon raqamingizni yuboring 👇",
        "send_phone_button": "📞 Raqamni yuborish",
        "send_phone_again": "❗ Iltimos, pastdagi tugma orqali telefon raqamingizni yuboring 👇",
        "registration_done": "✅ Ro‘yxatdan o‘tish yakunlandi.\n\nQuyidagilardan birini tanlang 👇",
        "course_btn": "🎓 O‘quv kursi haqida ma’lumot",
        "service_btn": "🔧 Servis haqida ma’lumot",
        "enter_job": "🧑‍💼 Kasbingizni kiriting:",
        "enter_skill": "🚗 Mashina servisi sohasida tajribangiz bormi?",
        "yes": "✅ Ha",
        "no": "❌ Yo‘q",
        "only_buttons": "❗ Iltimos, faqat quyidagi tugmalardan birini tanlang 👇",
        "skill_saved": "✅ Javob qabul qilindi.",
        "back_to_menu": "🔙 Siz asosiy menyuga qaytdingiz.",
        "skill_accepted": "✅ Javob qabul qilindi.",
        "call_btn": "📞 Bizga qo‘ng‘iroq qilish",
        "telegram_btn": "💬 Telegram orqali bog‘lanish",
        "instagram_btn": "📸 Bizga obuna bo‘ling",
        "back_to_menu": "⬅️ Asosiy menyuga qaytish",
        "choose_action": "Quyidagilardan birini tanlang 👇",
        "thanks_contact": "☎️ Rahmat! Tez orada siz bilan aloqaga chiqamiz.",
        "new_contact_request": "📞 Savolingiz bo'lsa, biz sizga aloqaga chiqamiz",
        "course_info": (
            "⚡️ Zamonaviy kasb – elektromobil ta’miri!\n"
            "Bugun elektromobillar soni tez o‘smoqda, lekin ularni ta’mirlay oladigan mutaxassislar juda kam!\n"
            "Bizning kursda - Malakali ustozlar, real amaliy mashg‘ulotlar va zamonaviy uskunalar sizni kutmoqda.\n"
            "🎓 Darslar: nazariya + amaliyot\n"
            "🚗 Elektromobil va gibridlarning tuzilishini o‘rganasiz\n"
            "🔋 Batareyalar va elektr tizimlar bilan xavfsiz ishlashni bilib olasiz\n"
            "🧰 Amaliy mashg‘ulotlarda haqiqiy avtomobillar bilan ishlaysiz\n\n"
            "📅 Kurs davomiyligi: 6 oy\n"
            "💰 Narx: 7 100 000 So'm (Bir oy uchun)\n"
            "1-Yillik bo'lib to'lash imkoni bor\n\n"
            "📍 Manzil: Toshkent sh. Katta Ka'ni ko'chasi 3А\n"
            "☎️ +998992201111"
        ),
        "service_info": (
            "🚗 <b>Zamonaviy avtomobilga zamonaviy servis</b>\n\n"
            "⚡️ Biz zamonaviy elektromobillar va gibrid avtomobillarni ta’mirlash, diagnostika va texnik xizmat ko‘rsatishga ixtisoslashganmiz.\n\n"
            "🔧 <b>Xizmatlarimiz:</b>\n"
            "• Elektromobil batareyalarini tekshirish va ta’mirlash\n"
            "• Elektr tizimlari diagnostikasi\n"
            "• Gibrid tizimlarni texnik xizmat ko‘rsatish\n"
            "• Dvigatel, transmissiya\n"
            "• Dasturiy ta’minot yangilanishi va sozlash\n"
            "• Avtohalokatga uchragan mashinalarni tiklash\n\n"
            "💡 <b>Nega bizni tanlashadi?</b>\n"
            "✅ Malakali ustalar\n"
            "✅ Zamonaviy texnika va uskunalar\n"
            "✅ Sifatli ehtiyot qismlar\n"
            "✅ Buyurtma bo‘lsa chet eldan tez yetkazamiz\n"
            "✅ Ishonchli kafolat\n\n"
            "📍 <b>Manzil:</b> Toshkent, Katta Ka’ni ko‘chasi 3A\n"
            "☎️ <b>Aloqa:</b> +998 99 220 11 11\n\n"
            "🔋 Elektromobilingizni ishonchli qo‘llarga topshiring!"
        )
    },

    "🇷🇺 Русский": {
        "welcome": "🐼 <b>Добро пожаловать в Panda Service!</b>\n\nДля регистрации введите данные 👇",
        "choose_language": "Выберите язык 👇",
        "enter_name": "🧾 Введите ваше имя:",
        "enter_phone": "📱 Отправьте ваш номер телефона 👇",
        "send_phone_button": "📞 Отправить номер",
        "send_phone_again": "❗ Пожалуйста, отправьте номер через кнопку 👇",
        "registration_done": "✅ Регистрация завершена.\n\nВыберите раздел 👇",
        "course_btn": "🎓 Информация о курсе",
        "service_btn": "🔧 Информация о сервисе",
        "enter_job": "🧑‍💼 Введите вашу профессию:",
        "enter_skill": "🚗 Есть ли у вас опыт в сфере автомобильного сервиса?",
        "yes": "✅ Да",
        "no": "❌ Нет",
        "only_buttons": "❗ Пожалуйста, выберите только через кнопки 👇",
        "skill_saved": "✅ Ответ сохранён.",
        "back_to_menu": "🔙 Вы вернулись в главное меню.",
        "skill_accepted": "✅ Ответ принят.",
        "call_btn": "📞 Позвонить нам",
        "telegram_btn": "💬 Связаться через Telegram",
        "instagram_btn": "📸 Подписаться на Instagram",
        "back_to_menu": "⬅️ Вернуться в главное меню",
        "thanks_contact": "☎️ Спасибо! Мы скоро с вами свяжемся.",
        "choose_action": "Выберите действие 👇",
        "new_contact_request": "📞 Новый запрос на связь",
        "course_info": (
            "⚡️ Современная профессия – ремонт электромобилей!\n"
            "Сегодня количество электромобилей быстро растет, но специалистов по их ремонту очень мало!\n"
            "На нашем курсе – опытные наставники, реальные практические занятия и современное оборудование.\n"
            "🎓 Уроки: теория + практика\n"
            "🚗 Изучите устройство электромобилей и гибридов\n"
            "🔋 Научитесь безопасно работать с батареями и электрическими системами\n"
            "🧰 Практика на настоящих автомобилях\n\n"
            "📅 Продолжительность курса: 6 месяцев\n"
            "💰 Стоимость: 7 100 000 сум (в месяц)\n"
            "Возможна оплата 1 год по частям\n\n"
            "📍 Адрес: Ташкент, ул. Katta Ka’ni 3А\n"
            "☎️ +998992201111"
        ),
        "service_info": (
            "🚗 <b>Современный сервис для современных автомобилей</b>\n\n"
            "⚡️ Мы специализируемся на ремонте, диагностике и техническом обслуживании современных электромобилей и гибридов.\n\n"
            "🔧 <b>Наши услуги:</b>\n"
            "• Проверка и ремонт батарей электромобилей\n"
            "• Диагностика электрических систем\n"
            "• Обслуживание гибридных систем\n"
            "• Двигатель, трансмиссия\n"
            "• Обновление и настройка программного обеспечения\n"
            "• Восстановление автомобилей после аварий\n\n"
            "💡 <b>Почему выбирают нас?</b>\n"
            "✅ Квалифицированные мастера\n"
            "✅ Современная техника и оборудование\n"
            "✅ Качественные запчасти\n"
            "✅ Быстрая доставка при заказе из-за границы\n"
            "✅ Надёжная гарантия\n\n"
            "📍 <b>Адрес:</b> Ташкент, ул. Katta Ka’ni 3A\n"
            "☎️ <b>Контакт:</b> +998 99 220 11 11\n\n"
            "🔋 Доверьте свой электромобиль надежным рукам!"
        )
    }
}




# --- Функция для отправки в несколько групп ---
async def send_to_groups(bot, text: str):
    if not GROUP_IDS:
        print("⚠️ Список GROUP_IDS пустой. Проверьте переменные окружения.")
        return

    for group_id in GROUP_IDS:
        try:
            await bot.send_message(group_id, text, parse_mode="HTML")
            print(f"✅ Сообщение успешно отправлено в группу {group_id}")
        except Exception as e:
            # Разбираем основные ошибки
            if "bot was blocked by the user" in str(e):
                print(f"⚠️ Бот заблокирован в группе {group_id}")
            elif "Forbidden" in str(e):
                print(f"⚠️ Бот не является участником группы {group_id} или нет прав на отправку")
            elif "chat not found" in str(e):
                print(f"⚠️ Группа {group_id} не найдена. Проверьте правильность ID")
            else:
                print(f"⚠️ Ошибка при отправке в группу {group_id}: {e}")


# --- Excel fayl tayyorlash ---
def setup_excel():
    if not Path(EXCEL_FILE).exists():
        wb = Workbook()
        ws = wb.active
        ws.title = "Foydalanuvchilar"
        ws.append([
            "Ism", "Telefon", "Kategoriya", "Ko‘nikma",
            "Holat", "UserID", "Username", "Sana"
        ])
        wb.save(EXCEL_FILE)

setup_excel()

# --- FSM holatlar ---
class Registration(StatesGroup):
    choose_language = State()
    name = State()
    phone = State()
    category = State()
    skill = State()
    finish = State()


# --- /start ---
@router.message(CommandStart())
async def start(message: Message, state: FSMContext):
    await state.clear()

    try:
        photo = FSInputFile("logo.png")
        await message.answer_photo(photo, caption="🐼 Panda Servis!", parse_mode="HTML")
    except:
        await message.answer("🐼 Panda Servis!")

    lang_kb = ReplyKeyboardMarkup(
    resize_keyboard=True,
    keyboard=[
        [
            KeyboardButton(text="🇺🇿 O‘zbek"),
            KeyboardButton(text="🇷🇺 Русский")
        ]
    ]
    )


    await message.answer("Tilni tanlang / Выберите язык 👇", reply_markup=lang_kb)
    await state.set_state(Registration.choose_language)


#uzim qushkanim
@router.message(Registration.choose_language)
async def choose_language(message: Message, state: FSMContext):

    lang = message.text.strip()

    # Клавиатура выбора языка
    lang_kb = ReplyKeyboardMarkup(
        resize_keyboard=True,
        keyboard=[
            [
                KeyboardButton(text="🇺🇿 O‘zbek"),
                KeyboardButton(text="🇷🇺 Русский")
            ]
        ]
    )

    # Если выбран неправильный язык → повторно показываем кнопки
    if lang not in ["🇺🇿 O‘zbek", "🇷🇺 Русский"]:
        await message.answer(
            "❗ Tilni tugma orqali tanlang / Выберите язык через кнопку 👇",
            reply_markup=lang_kb
        )
        return

    await state.update_data(language=lang)
    await message.answer(LANG[lang]["enter_name"], reply_markup=ReplyKeyboardRemove())
    await state.set_state(Registration.name)





# --- Ism / Имя ---
@router.message(Registration.name)
async def process_name(message: Message, state: FSMContext):
    data = await state.get_data()
    lang = data.get("language", "🇺🇿 O‘zbek")  # <-- тильни оламиз

    await state.update_data(name=message.text)

    # Tugma matni ham lang bo‘yicha
    contact_btn = ReplyKeyboardMarkup(
        keyboard=[[KeyboardButton(text=LANG[lang]["send_phone_button"], request_contact=True)]],
        resize_keyboard=True,
        one_time_keyboard=True
    )

    await state.set_state(Registration.phone)
    await message.answer(LANG[lang]["enter_phone"], reply_markup=contact_btn)  # <-- matn ham lang bo‘yicha


# --- Telefon ---
@router.message(Registration.phone)
async def process_phone(message: Message, state: FSMContext):
    data = await state.get_data()
    lang = data.get("language", "🇺🇿 O‘zbek")

    if not message.contact:
        contact_btn = ReplyKeyboardMarkup(
            keyboard=[[KeyboardButton(text=LANG[lang]["send_phone_button"], request_contact=True)]],
            resize_keyboard=True,
            one_time_keyboard=True
        )
        await message.answer(LANG[lang]["send_phone_again"], reply_markup=contact_btn)
        return

    phone = message.contact.phone_number
    await state.update_data(phone=phone)
    user = message.from_user
    data = await state.get_data()

    # Excelga yozish
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    ws.append([
        data.get("name", "—"),
        phone,
        "—", "—", "—",
        "🆕 Yangi foydalanuvchi ro‘yxatdan o‘tdi",
        user.id,
        user.username or "—",
        datetime.now().strftime("%Y-%m-%d %H:%M")
    ])
    wb.save(EXCEL_FILE)
    wb.close()

    # Adminga yuborish
    if ADMIN_IDS:
        for admin_id in ADMIN_IDS:
            try:
                await message.bot.send_message(
                    admin_id,
                    f"🆕 <b>Yangi foydalanuvchi ro‘yxatdan o‘tdi!</b>\n\n"
                    f"🧾 <b>Ism:</b> {data.get('name', '—')}\n"
                    f"📱 <b>Telefon:</b> {phone}\n"
                    f"💬 <b>Username:</b> @{user.username if user.username else '—'}\n"
                    f"🆔 <b>ID:</b> {user.id}\n"
                    f"🕒 <b>Vaqt:</b> {datetime.now().strftime('%Y-%m-%d %H:%M')}",
                    parse_mode="HTML"
                )
            except Exception as e:
                print(f"⚠️ Admin {admin_id} ga yuborishda xato: {e}")

    # Guruhlarga yuborish
    text = (
        f"🆕 <b>Yangi foydalanuvchi ro‘yxatdan o‘tdi!</b>\n\n"
        f"🧾 <b>Ism:</b> {data.get('name', '—')}\n"
        f"📱 <b>Telefon:</b> {phone}\n"
        f"💬 <b>Username:</b> @{user.username if user.username else '—'}\n"
        f"🆔 <b>ID:</b> {user.id}\n"
        f"🕒 <b>Vaqt:</b> {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    )
    await send_to_groups(message.bot, text)

    # Tilga mos menyu
    menu = ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text=LANG[lang]["course_btn"])],
            [KeyboardButton(text=LANG[lang]["service_btn"])]
        ],
        resize_keyboard=True
    )
    await message.answer(LANG[lang]["registration_done"], reply_markup=menu)
    await state.set_state(Registration.category)


# --- Tanlov: O‘quv kursi / Servis ---
@router.message(Registration.category)
async def process_category(message: Message, state: FSMContext):
    data = await state.get_data()
    lang = data.get("language", "🇺🇿 O‘zbek")

    text = message.text.strip()

    valid_options = [
        LANG[lang]["course_btn"],
        LANG[lang]["service_btn"]
    ]

    if text not in valid_options:
        menu = ReplyKeyboardMarkup(
            keyboard=[[KeyboardButton(text=opt)] for opt in valid_options],
            resize_keyboard=True
        )
        await message.answer(LANG[lang]["only_buttons"], reply_markup=menu)
        return

    await state.update_data(category=text)

    # === O‘quv kursi ===
    if text == LANG[lang]["course_btn"]:
        buttons = ReplyKeyboardMarkup(
            keyboard=[[KeyboardButton(text=LANG[lang]["yes"]),
                       KeyboardButton(text=LANG[lang]["no"])]],
            resize_keyboard=True
        )
        await message.answer(LANG[lang]["enter_skill"], reply_markup=buttons)
        await state.set_state(Registration.skill)
        return

    # === Servis ===
    if text == LANG[lang]["service_btn"]:
        try:
            # FSM переходим в finish
            await state.set_state(Registration.finish)

            # Убираем ReplyKeyboard полностью
            await message.answer(
                "📌 Информация о сервисе:",
                reply_markup=ReplyKeyboardRemove()
            )

            # Отправка фото с описанием сервиса
            photo = FSInputFile("service_photo.jpg")
            await message.answer_photo(
                photo,
                caption=LANG[lang]["service_info"],
                parse_mode="HTML",
                reply_markup=None  # главное меню не показываем
            )

            # Отправка локации
            await message.answer_location(latitude=41.270408, longitude=69.171301)

            # Inline-кнопки
            ikb = InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text=LANG[lang]["call_btn"], callback_data="show_phone")],
                [InlineKeyboardButton(text=LANG[lang]["telegram_btn"], url="https://t.me/PandaAvtoservis")],
                [InlineKeyboardButton(text=LANG[lang]["instagram_btn"], url="https://www.instagram.com/panda.avtoservis")],
                [InlineKeyboardButton(text=LANG[lang]["back_to_menu"], callback_data="menu_back")]
            ])
            await message.answer(LANG[lang]["choose_action"], reply_markup=ikb)

        except Exception as e:
            print("⚠️ Servis bloki xatosi:", e)
   



# --- Ko‘nikma ---
@router.message(Registration.skill)
async def process_skill(message: Message, state: FSMContext):
    text = message.text.strip()

    data = await state.get_data()
    lang = data.get("language", "🇺🇿 O‘zbek")

    # Если ответ не соответствует кнопкам → повторно показать кнопки
    if text not in [LANG[lang]["yes"], LANG[lang]["no"]]:
        buttons = ReplyKeyboardMarkup(
            keyboard=[[KeyboardButton(text=LANG[lang]["yes"]), KeyboardButton(text=LANG[lang]["no"])]],
            resize_keyboard=True
        )
        await message.answer(LANG[lang]["only_buttons"], reply_markup=buttons)
        return

    # Сохраняем ответ
    await state.update_data(skill=text)
    await message.answer(LANG[lang]["skill_accepted"], reply_markup=ReplyKeyboardRemove())

    # --- Текст курса по выбранному языку ---
    info_text = LANG[lang]["course_info"]

    try:
        await message.answer_photo(FSInputFile("course_info.jpg"), caption=info_text, parse_mode="HTML")
    except Exception:
        await message.answer(info_text, parse_mode="HTML")

    # Отправляем местоположение
    await message.answer_location(latitude=41.270408, longitude=69.171301)

    # --- Инлайн-кнопки ---
    TELEGRAM_MANAGER = "PandaAvtoservis"
    ikb = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text=LANG[lang]["new_contact_request"], callback_data="call_request")],
        [InlineKeyboardButton(text=LANG[lang]["call_btn"], callback_data="show_phone")],
        [InlineKeyboardButton(text=LANG[lang]["telegram_btn"], url="https://t.me/PandaAvtoservis")],
        [InlineKeyboardButton(text=LANG[lang]["instagram_btn"], url="https://www.instagram.com/panda.avtoservis?igsh=cHFnNXRxOTY2aGlv")],
        [InlineKeyboardButton(text=LANG[lang]["back_to_menu"], callback_data="menu_back")]
    ])

    await message.answer(LANG[lang]["choose_action"], reply_markup=ikb)

    # FSMni finish holatiga o'tkazamiz
    await state.set_state(Registration.finish)





@router.callback_query(F.data == "kurs_yozilish")
async def callback_kurs_yozilish(callback: CallbackQuery, state: FSMContext):
    data = await state.get_data()
    user = callback.from_user

    # --- Excelga yozish ---
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    ws.append([
        data.get("name", "—"),
        data.get("phone", "—"),
        data.get("category", "—"),
        data.get("skill", "—"),
        "✅ Kursga yozildi",
        user.id,
        user.username or "—",
        datetime.now().strftime("%Y-%m-%d %H:%M")
    ])
    wb.save(EXCEL_FILE)
    wb.close()

    # --- Adminga yuborish ---
    if ADMIN_IDS:
        for admin_id in ADMIN_IDS:
            try:
                await callback.bot.send_message(
                    admin_id,
                    f"📚 <b>Yangi kursga yozilish!</b>\n\n"
                    f"🧾 Ism: {data.get('name', '—')}\n"
                    f"📱 Telefon: {data.get('phone', '—')}\n"
                    f"📚 Kategoriya: {data.get('category', '—')}\n"
                    f"⚙️ Ko‘nikma: {data.get('skill', '—')}\n"
                    f"💬 Username: @{user.username if user.username else '—'}\n"
                    f"🆔 ID: {user.id}\n"
                    f"🕒 {datetime.now().strftime('%Y-%m-%d %H:%M')}",
                    parse_mode="HTML"
                )
            except Exception as e:
                print(f"⚠️ Admin {admin_id} ga yuborishda xato: {e}")

# --- Отправка в группу ---
    text_group = (
        f"📚 <b>Yangi kursga yozilish!</b>\n\n"
        f"🧾 Ism: {data.get('name', '—')}\n"
        f"📱 Telefon: {data.get('phone', '—')}\n"
        f"📚 Kategoriya: {data.get('category', '—')}\n"
        f"⚙️ Ko‘nikma: {data.get('skill', '—')}\n"
        f"💬 Username: @{user.username if user.username else '—'}\n"
        f"🆔 ID: {user.id}\n"
        f"🕒 {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    )
    await send_to_groups(callback.bot, text_group)



    # --- Asosiy menyuga qaytish ---
    main_menu = ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text=LANG[lang]["course_btn"])],
            [KeyboardButton(text=LANG[lang]["service_btn"])]
        ],
        resize_keyboard=True
    )
    await callback.message.answer(LANG[lang]["back_to_menu"], reply_markup=main_menu)



# --- Qo‘ng‘iroq so‘rovi ---
# Пример для call_request
@router.callback_query(F.data == "call_request")
async def callback_call_request(callback: CallbackQuery, state: FSMContext):
    user = callback.from_user
    data = await state.get_data()
    phone = data.get("phone", "—")
    category = data.get("category", "🎓 O‘quv kursi haqida ma’lumot")
    skill = data.get("skill", "—")

    # Excel запись
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    ws.append([
        data.get("name", "—"),
        phone,
        category,
        skill,
        "📞 Qo‘ng‘iroq so‘rovi",
        user.id,
        user.username or "—",
        datetime.now().strftime("%Y-%m-%d %H:%M")
    ])
    wb.save(EXCEL_FILE)
    wb.close()

    # Пользователю
    data = await state.get_data()
    lang = data.get("language", "🇺🇿 O‘zbek")
    await callback.message.answer(LANG[lang]["thanks_contact"])


    # Админ уведомление
    if ADMIN_IDS:
        msg = (
            f"📞 <b>Yangi aloqa so‘rovi!</b>\n\n"
            f"🧾 <b>Ism:</b> {data.get('name', '—')}\n"
            f"📱 <b>Telefon:</b> {phone}\n"
            f"📚 <b>Kategoriya:</b> {category}\n"
            f"⚙️ <b>Ko‘nikma:</b> {skill}\n"
            f"💬 <b>Username:</b> @{user.username if user.username else '—'}\n"
            f"🆔 <b>ID:</b> {user.id}\n"
            f"🕒 <b>Vaqt:</b> {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        )
        for admin_id in ADMIN_IDS:
            try:
                await callback.bot.send_message(admin_id, msg, parse_mode="HTML")
            except Exception as e:
                print(f"⚠️ Admin {admin_id} ga yuborishda xato: {e}")

# --- Отправка в группу ---
    text_group = (
        f"📞 <b>Yangi aloqa so‘rovi!</b>\n\n"
        f"🧾 Ism: {data.get('name', '—')}\n"
        f"📱 Telefon: {phone}\n"
        f"📚 Kategoriya: {category}\n"
        f"⚙️ Ko‘nikma: {skill}\n"
        f"💬 Username: @{user.username if user.username else '—'}\n"
        f"🆔 ID: {user.id}\n"
        f"🕒 {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    )
    await send_to_groups(callback.bot, text_group)


    # Возврат в главное меню
    main_menu = ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text=LANG[lang]["course_btn"])],
            [KeyboardButton(text=LANG[lang]["service_btn"])]
        ],
        resize_keyboard=True
    )
    await callback.message.answer(LANG[lang]["back_to_menu"], reply_markup=main_menu)
    await state.set_state(Registration.category)
    await callback.answer("So‘rov yuborildi ✅")





# --- Servisdagi Qo‘ng‘iroq qilish tugmasi ---
@router.callback_query(F.data == "show_phone")
async def callback_show_phone(callback: CallbackQuery, state: FSMContext):
    user = callback.from_user
    data = await state.get_data()

    phone_number = "+998992201111"

    # 1. Foydalanuvchiga raqamni ko‘rsatamiz
    await callback.message.answer(f"<a href='tel:{phone_number}'>📱 {phone_number}</a>\n\n", parse_mode="HTML")
    await callback.answer("Raqam ko‘rsatildi ✅")

    # 2. Excel faylga yozamiz (faktochno faqat ism, telefon, kategoriya, job, skill)
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    ws.append([
        data.get("name", "—"),
        data.get("phone", "—"),
        data.get("category", "🔧 Servis haqida ma’lumot"),
        data.get("skill", "—"),
        "📞 Qo‘ng‘iroq tugmasi bosildi",
        user.id,
        user.username or "—",
        datetime.now().strftime("%Y-%m-%d %H:%M")
    ])
    wb.save(EXCEL_FILE)
    wb.close()

    # 3. Adminga yuboramiz
    if ADMIN_IDS:
        msg = (
            "📞 <b>Foydalanuvchi qo‘ng‘iroq qilish tugmasini bosdi!</b>\n\n"
            f"🧾 <b>Ism:</b> {data.get('name', '—')}\n"
            f"📱 <b>Telefon:</b> {data.get('phone', '—')}\n"
            f"📚 <b>Kategoriya:</b> {data.get('category', '🔧 Servis haqida ma’lumot')}\n"
            f"⚙️ <b>Ko‘nikma:</b> {data.get('skill', '—')}\n"
            f"💬 <b>Username:</b> @{user.username if user.username else '—'}\n"
            f"🆔 <b>ID:</b> {user.id}\n"
            f"🕒 <b>Vaqt:</b> {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        )
        for admin_id in ADMIN_IDS:
            try:
                await callback.bot.send_message(admin_id, msg, parse_mode="HTML")
            except Exception as e:
                print(f"⚠️ Admin {admin_id} ga yuborishda xato: {e}")


# --- Отправка в группу ---
    text_group = (
        f"📞 <b>Foydalanuvchi qo‘ng‘iroq tugmasini bosdi!</b>\n\n"
        f"🧾 Ism: {data.get('name', '—')}\n"
        f"📱 Telefon: {data.get('phone', '—')}\n"
        f"📚 Kategoriya: {data.get('category', '—')}\n"
        f"⚙️ Ko‘nikma: {data.get('skill', '—')}\n"
        f"💬 Username: @{user.username if user.username else '—'}\n"
        f"🆔 ID: {user.id}\n"
        f"🕒 {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    )
    await send_to_groups(callback.bot, text_group)







    

# --- Telegram orqali bog‘lanish (faqat Excel yozish + menejerga xabar yuborish) ---
@router.callback_query(F.data == "telegram_connect")
async def telegram_connect(callback: CallbackQuery, state: FSMContext):
    TELEGRAM_MANAGER = "PandaAvtoservis"
    data = await state.get_data()
    user = callback.from_user

    # Excel yozish faqat kerakli maydonlar bilan
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    ws.append([
        data.get("name", "—"),
        data.get("phone", "—"),
        data.get("category", "—"),
        data.get("skill", "—"),
        "💬 Telegram orqali bog‘landi",
        user.id,
        user.username or "—",
        datetime.now().strftime("%Y-%m-%d %H:%M")
    ])
    wb.save(EXCEL_FILE)
    wb.close()


    # --- Asosiy menyuga qaytish (ham kurs, ham servis uchun ishlaydi) ---
@router.callback_query(F.data == "menu_back")
async def callback_menu_back(callback: CallbackQuery, state: FSMContext):
    data = await state.get_data()
    lang = data.get("language", "🇺🇿 O‘zbek")  # получаем язык пользователя

    # Восстанавливаем состояние категории
    await state.set_state(Registration.category)

    # Основное меню с учетом языка
    main_menu = ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text=LANG[lang]["course_btn"])],
            [KeyboardButton(text=LANG[lang]["service_btn"])]
        ],
        resize_keyboard=True
    )

    await callback.message.answer(LANG[lang]["back_to_menu"], reply_markup=main_menu)
    await callback.answer()



   


    # --- Faqat yo‘naltirish tugmasi ---
    ikb = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(
            text="💬 Menejer bilan bog‘lanish",
            url=f"https://t.me/{TELEGRAM_MANAGER}"
        )]
    ])

    # 🔇 Foydalanuvchiga hech qanday xabar yubormaymiz — faqat yo‘naltirish
    await callback.answer()  # убирает "loading" кружок
    await callback.message.edit_reply_markup(reply_markup=ikb)

    # ❗ Запрещаем писать текст в меню курса (только инлайн кнопки!)
@router.message(Registration.finish)
async def block_text_in_course_menu(message: Message, state: FSMContext):
    data = await state.get_data()
    lang = data.get("language", "🇺🇿 O‘zbek")

    await message.answer(LANG[lang]["only_buttons"])







   

